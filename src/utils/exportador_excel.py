# src/utils/exportador_excel.py
"""
📊 EXPORTADOR A EXCEL - VERSIÓN COMPLETA
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import calendar
from pathlib import Path

class ExportadorExcel:
    """Exporta reportes financieros a Excel con formato profesional."""

    COLORES = {
        'primario': '667eea',
        'secundario': '764ba2',
        'exito': '00b894',
        'peligro': 'ff6b6b',
        'info': '74b9ff',
        'blanco': 'ffffff',
        'texto': '2d3436',
        'borde': 'dfe6e9'
    }

    @staticmethod
    def exportar_reporte_mensual(datos, mes, año, nombre_archivo=None):
        """Genera un archivo Excel con el reporte mensual completo."""
        
        if not nombre_archivo:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"reporte_{año}_{mes:02d}_{timestamp}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # 1. Hoja de Resumen
        ExportadorExcel._crear_hoja_resumen(wb, datos, mes, año)
        
        # 2. Hoja de Boletas
        ExportadorExcel._crear_hoja_boletas(wb, datos.get('boletas', []))
        
        # 3. Hoja de Facturas
        ExportadorExcel._crear_hoja_facturas(wb, datos.get('facturas', []))
        
        # 4. Hoja de Impuestos
        ExportadorExcel._crear_hoja_impuestos(wb, datos, mes, año)
        
        # Guardar en la raíz del proyecto
        ruta_archivo = Path(nombre_archivo)
        wb.save(ruta_archivo)
        
        return str(ruta_archivo)

    @staticmethod
    def _crear_hoja_resumen(wb, datos, mes, año):
        """Crea la hoja de resumen."""
        ws = wb.create_sheet("📊 Resumen")
        
        titulo_font = Font(name='Arial', size=18, bold=True, color='ffffff')
        titulo_fill = PatternFill(start_color=ExportadorExcel.COLORES['primario'], 
                                  end_color=ExportadorExcel.COLORES['primario'], 
                                  fill_type='solid')
        borde = Border(
            left=Side(style='thin', color=ExportadorExcel.COLORES['borde']),
            right=Side(style='thin', color=ExportadorExcel.COLORES['borde']),
            top=Side(style='thin', color=ExportadorExcel.COLORES['borde']),
            bottom=Side(style='thin', color=ExportadorExcel.COLORES['borde'])
        )
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f"📊 REPORTE FINANCIERO - {calendar.month_name[mes]} {año}"
        ws['A1'].font = titulo_font
        ws['A1'].fill = titulo_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(name='Arial', size=9, color=ExportadorExcel.COLORES['texto'])
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.row_dimensions[3].height = 10
        
        metricas = [
            ('💰 Ventas Totales', f"S/ {datos.get('total_ventas', 0):,.2f}"),
            ('📦 Compras Totales', f"S/ {datos.get('total_compras', 0):,.2f}"),
            ('💵 Utilidad', f"S/ {datos.get('utilidad', 0):,.2f}"),
            ('💸 Impuesto RUS', f"S/ {datos.get('impuesto', 0):,.2f}"),
            ('⭐ Percepciones', f"S/ {datos.get('total_percepciones', 0):,.2f}"),
        ]
        
        fila = 5
        for i, (etiqueta, valor) in enumerate(metricas):
            col = (i % 3) * 2 + 1
            fila_actual = fila + (i // 3) * 2
            
            celda_etiqueta = ws.cell(row=fila_actual, column=col)
            celda_etiqueta.value = etiqueta
            celda_etiqueta.alignment = Alignment(horizontal='center')
            celda_etiqueta.font = Font(name='Arial', size=10, color=ExportadorExcel.COLORES['texto'])
            
            celda_valor = ws.cell(row=fila_actual + 1, column=col)
            celda_valor.value = valor
            celda_valor.alignment = Alignment(horizontal='center')
            celda_valor.font = Font(name='Arial', size=14, bold=True, color=ExportadorExcel.COLORES['texto'])
            
            ws.column_dimensions[chr(64 + col)].width = 20
        
        fila_inicio_detalle = fila + 4
        ws.merge_cells(f'A{fila_inicio_detalle}:C{fila_inicio_detalle}')
        ws[f'A{fila_inicio_detalle}'] = "📋 RESUMEN DE DOCUMENTOS"
        ws[f'A{fila_inicio_detalle}'].font = Font(name='Arial', size=12, bold=True, color=ExportadorExcel.COLORES['texto'])
        
        fila_detalle = fila_inicio_detalle + 2
        
        resumen_data = [
            ['TIPO', 'CANTIDAD', 'MONTO TOTAL'],
            ['Boletas de Venta', len(datos.get('boletas', [])), f"S/ {datos.get('total_ventas', 0):,.2f}"],
            ['Facturas de Compra', len(datos.get('facturas', [])), f"S/ {datos.get('total_compras', 0):,.2f}"],
            ['Percepciones', len([f for f in datos.get('facturas', []) if f.get('percepcion', 0) > 0]), f"S/ {datos.get('total_percepciones', 0):,.2f}"],
        ]
        
        for i, row in enumerate(resumen_data):
            for j, value in enumerate(row):
                celda = ws.cell(row=fila_detalle + i, column=j + 1)
                celda.value = value
                celda.border = borde
                celda.alignment = Alignment(horizontal='center' if i == 0 else 'left' if j == 0 else 'right')
                
                if i == 0:
                    celda.font = Font(name='Arial', size=10, bold=True, color='ffffff')
                    celda.fill = PatternFill(start_color=ExportadorExcel.COLORES['secundario'], 
                                            end_color=ExportadorExcel.COLORES['secundario'], 
                                            fill_type='solid')
                else:
                    celda.font = Font(name='Arial', size=10, color=ExportadorExcel.COLORES['texto'])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20

    @staticmethod
    def _crear_hoja_boletas(wb, boletas):
        """Crea la hoja de boletas."""
        ws = wb.create_sheet("📄 Boletas")
        
        if not boletas:
            ws['A1'] = "No hay boletas registradas"
            return
        
        # Crear DataFrame con datos seguros
        datos_boletas = []
        for b in boletas:
            if hasattr(b, 'numero_boleta'):  # Objeto SQLAlchemy
                datos_boletas.append({
                    'numero': b.numero_boleta or 'N/A',
                    'fecha': b.fecha_emision.strftime('%d/%m/%Y') if b.fecha_emision else 'N/A',
                    'cliente': b.nombre_cliente or 'Cliente General',
                    'monto': b.monto_total or 0
                })
            else:  # Diccionario
                datos_boletas.append({
                    'numero': b.get('numero', b.get('numero_boleta', 'N/A')),
                    'fecha': b.get('fecha_emision_str', b.get('fecha', 'N/A')),
                    'cliente': b.get('nombre_cliente', b.get('cliente', 'Cliente General')),
                    'monto': b.get('monto_total', b.get('monto', 0))
                })
        
        df = pd.DataFrame(datos_boletas)
        
        # Verificar y crear columnas si no existen
        columnas_requeridas = ['numero', 'fecha', 'cliente', 'monto']
        for col in columnas_requeridas:
            if col not in df.columns:
                df[col] = 'N/A'
        
        df_display = df[columnas_requeridas].copy()
        
        # Formatear montos
        df_display['monto'] = df_display['monto'].apply(
            lambda x: f"S/ {float(x):,.2f}" if isinstance(x, (int, float)) else x
        )
        
        df_display.columns = ['Número', 'Fecha', 'Cliente', 'Monto']
        
        for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        cabecera_font = Font(name='Arial', size=10, bold=True, color='ffffff')
        cabecera_fill = PatternFill(start_color=ExportadorExcel.COLORES['primario'], 
                                   end_color=ExportadorExcel.COLORES['primario'], 
                                   fill_type='solid')
        
        for col in range(1, 5):
            celda = ws.cell(row=1, column=col)
            celda.font = cabecera_font
            celda.fill = cabecera_fill
            celda.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20

    @staticmethod
    def _crear_hoja_facturas(wb, facturas):
        """Crea la hoja de facturas."""
        ws = wb.create_sheet("📄 Facturas")
        
        if not facturas:
            ws['A1'] = "No hay facturas registradas"
            return
        
        datos_facturas = []
        for f in facturas:
            if hasattr(f, 'numero_factura'):  # Objeto SQLAlchemy
                datos_facturas.append({
                    'numero': f.numero_factura or 'N/A',
                    'fecha': f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else 'N/A',
                    'proveedor': f.proveedor or 'N/A',
                    'sub_total': f.sub_total or 0,
                    'igv': f.igv or 0,
                    'total': f.total_pagar or 0,
                    'percepcion': f.percepcion or 0
                })
            else:  # Diccionario
                datos_facturas.append({
                    'numero': f.get('numero', f.get('numero_factura', 'N/A')),
                    'fecha': f.get('fecha_emision_str', f.get('fecha', 'N/A')),
                    'proveedor': f.get('proveedor', 'N/A'),
                    'sub_total': f.get('sub_total', 0),
                    'igv': f.get('igv', 0),
                    'total': f.get('total_pagar', f.get('total', 0)),
                    'percepcion': f.get('percepcion', 0)
                })
        
        df = pd.DataFrame(datos_facturas)
        
        columnas_requeridas = ['numero', 'fecha', 'proveedor', 'sub_total', 'igv', 'total', 'percepcion']
        for col in columnas_requeridas:
            if col not in df.columns:
                df[col] = 'N/A'
        
        df_display = df[columnas_requeridas].copy()
        
        for col in ['sub_total', 'igv', 'total', 'percepcion']:
            df_display[col] = df_display[col].apply(
                lambda x: f"S/ {float(x):,.2f}" if isinstance(x, (int, float)) else x
            )
        
        df_display.columns = ['Número', 'Fecha', 'Proveedor', 'Sub Total', 'IGV', 'Total', 'Percepción']
        
        for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        cabecera_font = Font(name='Arial', size=10, bold=True, color='ffffff')
        cabecera_fill = PatternFill(start_color=ExportadorExcel.COLORES['secundario'], 
                                   end_color=ExportadorExcel.COLORES['secundario'], 
                                   fill_type='solid')
        
        for col in range(1, 8):
            celda = ws.cell(row=1, column=col)
            celda.font = cabecera_font
            celda.fill = cabecera_fill
            celda.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18

    @staticmethod
    def _crear_hoja_impuestos(wb, datos, mes, año):
        """Crea la hoja de impuestos."""
        ws = wb.create_sheet("💰 Impuestos")
        
        titulo_font = Font(name='Arial', size=14, bold=True, color=ExportadorExcel.COLORES['primario'])
        
        ws.merge_cells('A1:C1')
        ws['A1'] = f"💰 CÁLCULO DE IMPUESTOS - {calendar.month_name[mes]} {año}"
        ws['A1'].font = titulo_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        total_ventas = datos.get('total_ventas', 0)
        total_percepciones = datos.get('total_percepciones', 0)
        impuesto = datos.get('impuesto', 0)
        
        fila = 3
        calculos = [
            ('📊 Ventas Totales', total_ventas),
            ('📊 Percepciones', total_percepciones),
            ('', ''),
            ('📊 Tasa RUS', '5%'),
            ('📊 Impuesto Base', total_ventas * 0.05),
            ('⭐ Percepciones', total_percepciones),
            ('', ''),
            ('✅ IMPUESTO FINAL', impuesto),
        ]
        
        for i, (etiqueta, valor) in enumerate(calculos):
            fila_actual = fila + i
            
            if not etiqueta:
                continue
            
            celda_etiqueta = ws.cell(row=fila_actual, column=1)
            celda_etiqueta.value = etiqueta
            celda_etiqueta.font = Font(name='Arial', size=11, color=ExportadorExcel.COLORES['texto'])
            
            celda_valor = ws.cell(row=fila_actual, column=2)
            if isinstance(valor, str):
                celda_valor.value = valor
            else:
                celda_valor.value = f"S/ {valor:,.2f}"
            
            celda_valor.alignment = Alignment(horizontal='right')
            
            if 'FINAL' in etiqueta:
                celda_valor.font = Font(name='Arial', size=14, bold=True, color='ffffff')
                celda_valor.fill = PatternFill(start_color=ExportadorExcel.COLORES['exito'], 
                                              end_color=ExportadorExcel.COLORES['exito'], 
                                              fill_type='solid')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25