"""
Servicio de Exportación a Excel
Genera reportes profesionales en formato Excel
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import db
from src.models.factura_compra import FacturaCompra
from src.models.boleta_venta import BoletaVenta
from src.models.percepcion import Percepcion


class ExportService:
    """Servicio para exportar datos a Excel"""
    
    @staticmethod
    def exportar_reporte_completo(mes=None, anio=None):
        """Exporta un reporte completo a Excel con múltiples hojas"""
        
        facturas = FacturaCompra.query.all()
        boletas = BoletaVenta.query.all()
        percepciones = Percepcion.query.all()
        
        if mes and anio:
            facturas = [f for f in facturas if f.mes == mes and f.anio == anio]
            boletas = [b for b in boletas if b.mes == mes and b.anio == anio]
            percepciones = [p for p in percepciones if p.mes == mes and p.anio == anio]
        elif anio:
            facturas = [f for f in facturas if f.anio == anio]
            boletas = [b for b in boletas if b.anio == anio]
            percepciones = [p for p in percepciones if p.anio == anio]
        elif mes:
            facturas = [f for f in facturas if f.mes == mes]
            boletas = [b for b in boletas if b.mes == mes]
            percepciones = [p for p in percepciones if p.mes == mes]
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
        title_font = Font(name='Arial', size=16, bold=True, color='1a3a5c')
        value_font = Font(name='Arial', size=14, bold=True)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # HOJA 1: RESUMEN
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"
        
        ws_resumen.merge_cells('A1:D1')
        ws_resumen['A1'] = '📊 REPORTE FINANCIERO RUS'
        ws_resumen['A1'].font = title_font
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        
        ws_resumen.merge_cells('A2:D2')
        ws_resumen['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen['A2'].font = Font(name='Arial', size=10, italic=True, color='888888')
        ws_resumen['A2'].alignment = Alignment(horizontal='center')
        
        filtros_texto = "Todos"
        if mes and anio:
            filtros_texto = f"Mes: {mes}, Año: {anio}"
        elif anio:
            filtros_texto = f"Año: {anio}"
        elif mes:
            filtros_texto = f"Mes: {mes}"
        
        ws_resumen.merge_cells('A3:D3')
        ws_resumen['A3'] = f'Filtros: {filtros_texto}'
        ws_resumen['A3'].font = Font(name='Arial', size=10, color='666666')
        ws_resumen['A3'].alignment = Alignment(horizontal='center')
        
        ws_resumen.row_dimensions[4].height = 15
        
        headers = ['Métrica', 'Valor', 'Detalle', '']
        for col, header in enumerate(headers, 1):
            cell = ws_resumen.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        total_ventas = sum(b.monto for b in boletas)
        total_compras = sum(f.monto for f in facturas)
        total_percepciones = sum(p.monto for p in percepciones)
        utilidad = total_ventas - total_compras
        total_documentos = len(facturas) + len(boletas) + len(percepciones)
        
        datos_resumen = [
            ['💰 Ventas Totales', f'S/ {total_ventas:,.2f}', f'{len(boletas)} boletas', ''],
            ['🛒 Compras Totales', f'S/ {total_compras:,.2f}', f'{len(facturas)} facturas', ''],
            ['📋 Percepciones', f'S/ {total_percepciones:,.2f}', f'{len(percepciones)} documentos', ''],
            ['📈 Utilidad', f'S/ {utilidad:,.2f}', f'{((utilidad/total_ventas)*100) if total_ventas > 0 else 0:.1f}% margen', ''],
            ['📄 Total Documentos', str(total_documentos), 'documentos procesados', ''],
        ]
        
        for row_idx, row_data in enumerate(datos_resumen, 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center')
                if col_idx == 2:
                    cell.font = value_font
        
        for col in range(1, 4):
            ws_resumen.column_dimensions[get_column_letter(col)].width = 25
        
        # HOJA 2: FACTURAS
        ws_facturas = wb.create_sheet("Facturas")
        headers_facturas = ['Número', 'Fecha', 'Monto', 'Impuesto', 'Proveedor', 'Descripción']
        for col, header in enumerate(headers_facturas, 1):
            cell = ws_facturas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, f in enumerate(facturas, 2):
            ws_facturas.cell(row=row_idx, column=1, value=f.numero).border = border
            ws_facturas.cell(row=row_idx, column=2, value=f.fecha_emision.strftime('%d/%m/%Y')).border = border
            ws_facturas.cell(row=row_idx, column=3, value=f.monto).number_format = '"S/" #,##0.00'
            ws_facturas.cell(row=row_idx, column=3).border = border
            ws_facturas.cell(row=row_idx, column=4, value=f.impuesto).number_format = '"S/" #,##0.00'
            ws_facturas.cell(row=row_idx, column=4).border = border
            ws_facturas.cell(row=row_idx, column=5, value=f.proveedor or 'Natura Cosméticos S.A.').border = border
            ws_facturas.cell(row=row_idx, column=6, value=f.descripcion or '').border = border
        
        for col in range(1, 7):
            ws_facturas.column_dimensions[get_column_letter(col)].width = 20
        
        # HOJA 3: BOLETAS
        ws_boletas = wb.create_sheet("Boletas")
        headers_boletas = ['Número', 'Fecha', 'Monto', 'Cliente', 'Descripción']
        for col, header in enumerate(headers_boletas, 1):
            cell = ws_boletas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, b in enumerate(boletas, 2):
            ws_boletas.cell(row=row_idx, column=1, value=b.numero).border = border
            ws_boletas.cell(row=row_idx, column=2, value=b.fecha_emision.strftime('%d/%m/%Y')).border = border
            ws_boletas.cell(row=row_idx, column=3, value=b.monto).number_format = '"S/" #,##0.00'
            ws_boletas.cell(row=row_idx, column=3).border = border
            ws_boletas.cell(row=row_idx, column=4, value=b.cliente or 'Cliente no registrado').border = border
            ws_boletas.cell(row=row_idx, column=5, value=b.descripcion or '').border = border
        
        for col in range(1, 6):
            ws_boletas.column_dimensions[get_column_letter(col)].width = 20
        
        # HOJA 4: PERCEPCIONES
        ws_percepciones = wb.create_sheet("Percepciones")
        headers_percepciones = ['Número', 'Fecha', 'Monto', 'Proveedor', 'Descripción']
        for col, header in enumerate(headers_percepciones, 1):
            cell = ws_percepciones.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, p in enumerate(percepciones, 2):
            ws_percepciones.cell(row=row_idx, column=1, value=p.numero).border = border
            ws_percepciones.cell(row=row_idx, column=2, value=p.fecha_emision.strftime('%d/%m/%Y')).border = border
            ws_percepciones.cell(row=row_idx, column=3, value=p.monto).number_format = '"S/" #,##0.00'
            ws_percepciones.cell(row=row_idx, column=3).border = border
            ws_percepciones.cell(row=row_idx, column=4, value=p.proveedor or 'Natura Cosméticos S.A.').border = border
            ws_percepciones.cell(row=row_idx, column=5, value=p.descripcion or '').border = border
        
        for col in range(1, 6):
            ws_percepciones.column_dimensions[get_column_letter(col)].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output